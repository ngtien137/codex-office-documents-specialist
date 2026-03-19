#!/usr/bin/env python3
"""Inspect XLSX/XLSM files and emit a JSON summary."""

from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any, Dict, List

from openpyxl import load_workbook


def preview(value: Any, limit: int = 80) -> str:
    if value is None:
        return ""
    text = " ".join(str(value).split())
    if len(text) <= limit:
        return text
    return text[: limit - 3].rstrip() + "..."


def trimmed(values: List[str]) -> List[str]:
    items = list(values)
    while items and items[-1] == "":
        items.pop()
    return items


def sample_rows(worksheet, max_rows: int, max_cols: int) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    scan_limit = min(worksheet.max_row or 1, 200)
    for row_index in range(1, scan_limit + 1):
        values = []
        non_empty = False
        for column_index in range(1, min(worksheet.max_column or 1, max_cols) + 1):
            value = worksheet.cell(row=row_index, column=column_index).value
            if value not in (None, ""):
                non_empty = True
            values.append(preview(value))
        compact = trimmed(values)
        if non_empty and compact:
            rows.append({"row": row_index, "values": compact})
        if len(rows) >= max_rows:
            break
    return rows


def scan_cells(worksheet) -> Dict[str, Any]:
    estimated_cells = max(1, (worksheet.max_row or 1) * (worksheet.max_column or 1))
    full_scan = estimated_cells <= 50000
    max_rows = worksheet.max_row if full_scan else min(worksheet.max_row or 1, 500)
    max_cols = worksheet.max_column if full_scan else min(worksheet.max_column or 1, 50)

    formula_count = 0
    comment_count = 0
    hyperlink_count = 0
    non_empty_cell_count = 0

    for row in worksheet.iter_rows(min_row=1, max_row=max_rows, min_col=1, max_col=max_cols):
        for cell in row:
            if cell.value not in (None, ""):
                non_empty_cell_count += 1
            if isinstance(cell.value, str) and cell.value.startswith("="):
                formula_count += 1
            if cell.comment is not None:
                comment_count += 1
            if cell.hyperlink is not None:
                hyperlink_count += 1

    return {
        "scanned_entire_used_range": full_scan,
        "scan_max_rows": max_rows,
        "scan_max_columns": max_cols,
        "non_empty_cell_count": non_empty_cell_count,
        "formula_count": formula_count,
        "comment_count": comment_count,
        "hyperlink_count": hyperlink_count,
    }


def sheet_summary(worksheet, sample_row_limit: int, sample_col_limit: int) -> Dict[str, Any]:
    cell_scan = scan_cells(worksheet)
    table_names = sorted(list(worksheet.tables.keys()))
    merged_ranges = [str(item) for item in list(worksheet.merged_cells.ranges)[:12]]

    return {
        "title": worksheet.title,
        "sheet_state": worksheet.sheet_state,
        "max_row": worksheet.max_row,
        "max_column": worksheet.max_column,
        "dimensions": worksheet.calculate_dimension(),
        "table_names": table_names,
        "table_count": len(table_names),
        "chart_count": len(getattr(worksheet, "_charts", [])),
        "image_count": len(getattr(worksheet, "_images", [])),
        "merged_range_count": len(worksheet.merged_cells.ranges),
        "merged_ranges_preview": merged_ranges,
        "cell_scan": cell_scan,
        "sample_rows": sample_rows(worksheet, sample_row_limit, sample_col_limit),
    }


def workbook_properties(workbook) -> Dict[str, Any]:
    props = workbook.properties
    return {
        "title": props.title,
        "subject": props.subject,
        "creator": props.creator,
        "description": props.description,
        "keywords": props.keywords,
        "category": props.category,
        "created": props.created.isoformat() if props.created else None,
        "modified": props.modified.isoformat() if props.modified else None,
        "last_modified_by": props.lastModifiedBy,
    }


def build_report(path: Path, sample_row_limit: int, sample_col_limit: int) -> Dict[str, Any]:
    workbook = load_workbook(filename=path, data_only=False, keep_vba=True)
    named_ranges = sorted(list(workbook.defined_names.keys()))
    sheets = [sheet_summary(ws, sample_row_limit, sample_col_limit) for ws in workbook.worksheets]
    macro_suffixes = {".xlsm", ".xltm", ".xlam"}

    total_tables = sum(item["table_count"] for item in sheets)
    total_charts = sum(item["chart_count"] for item in sheets)
    total_images = sum(item["image_count"] for item in sheets)
    total_formulas = sum(item["cell_scan"]["formula_count"] for item in sheets)
    total_comments = sum(item["cell_scan"]["comment_count"] for item in sheets)

    return {
        "path": str(path.resolve()),
        "properties": workbook_properties(workbook),
        "summary": {
            "sheet_count": len(workbook.sheetnames),
            "named_range_count": len(named_ranges),
            "table_count": total_tables,
            "chart_count": total_charts,
            "image_count": total_images,
            "formula_count": total_formulas,
            "comment_count": total_comments,
            "has_vba": path.suffix.lower() in macro_suffixes,
        },
        "sheet_names": workbook.sheetnames,
        "named_ranges": named_ranges,
        "sheets": sheets,
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Inspect an XLSX or XLSM file and emit a JSON summary.")
    parser.add_argument("path", help="Path to the XLSX or XLSM file.")
    parser.add_argument("--sample-rows", type=int, default=6, help="Maximum non-empty rows to preview per sheet.")
    parser.add_argument("--sample-cols", type=int, default=8, help="Maximum columns to preview per sampled row.")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    report = build_report(Path(args.path), sample_row_limit=args.sample_rows, sample_col_limit=args.sample_cols)
    print(json.dumps(report, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
