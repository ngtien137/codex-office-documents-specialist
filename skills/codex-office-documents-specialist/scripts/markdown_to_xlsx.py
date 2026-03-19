#!/usr/bin/env python3
"""Create an XLSX workbook from Markdown headings and tables."""

from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

SHEET_HEADING_PATTERN = re.compile(r"^##\s+Sheet:\s+(.+)$")
TABLE_SEPARATOR_PATTERN = re.compile(r"^:?-{3,}:?$")

THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


def read_markdown(path: str | None, inline_markdown: str | None) -> str:
    if inline_markdown is not None:
        return inline_markdown
    if path is None:
        return sys.stdin.read()
    return Path(path).read_text(encoding="utf-8")


def split_table_row(line: str) -> List[str]:
    text = line.strip()
    if text.startswith("|"):
        text = text[1:]
    if text.endswith("|"):
        text = text[:-1]
    return [cell.replace("\\|", "|").strip() for cell in text.split("|")]


def is_separator_row(cells: List[str]) -> bool:
    return bool(cells) and all(TABLE_SEPARATOR_PATTERN.match(cell or "") for cell in cells)


def unique_headers(cells: List[str]) -> List[str]:
    headers: List[str] = []
    seen: dict[str, int] = {}
    for index, cell in enumerate(cells, start=1):
        base = cell or f"Column{index}"
        count = seen.get(base, 0) + 1
        seen[base] = count
        headers.append(base if count == 1 else f"{base}_{count}")
    return headers


def write_heading(worksheet, row_index: int, text: str, level: int) -> int:
    cell = worksheet.cell(row=row_index, column=1, value=text)
    size_map = {1: 18, 2: 16, 3: 14}
    cell.font = Font(size=size_map.get(level, 12), bold=True, color="1F1F1F")
    cell.alignment = Alignment(vertical="center")
    return row_index + 2


def write_paragraph(worksheet, row_index: int, text: str) -> int:
    cell = worksheet.cell(row=row_index, column=1, value=text)
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    return row_index + 1


def write_table(worksheet, row_index: int, table_rows: List[List[str]], table_number: int) -> tuple[int, int]:
    rows = [list(row) for row in table_rows]
    column_count = max(len(row) for row in rows)
    rows = [row + [""] * (column_count - len(row)) for row in rows]
    rows[0] = unique_headers(rows[0])

    for row_offset, row in enumerate(rows):
        for col_offset, value in enumerate(row):
            cell = worksheet.cell(row=row_index + row_offset, column=col_offset + 1)
            cell.value = value
            cell.border = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if row_offset == 0:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(fill_type="solid", fgColor="2F75B5")
            else:
                cell.font = Font(color="1F1F1F")

    end_row = row_index + len(rows) - 1
    end_col = column_count
    ref = f"A{row_index}:{get_column_letter(end_col)}{end_row}"
    table = Table(displayName=f"Table{table_number}", ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(table)
    return end_row + 2, end_col


def auto_fit_columns(worksheet, max_width: int = 40) -> None:
    for column_cells in worksheet.columns:
        max_length = 0
        for cell in column_cells:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        width = min(max(max_length + 2, 10), max_width)
        worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = width


def build_workbook(markdown: str, title: str | None, author: str | None, subject: str | None) -> tuple[Workbook, dict]:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"

    if title:
        workbook.properties.title = title
    if author:
        workbook.properties.creator = author
    if subject:
        workbook.properties.subject = subject

    lines = markdown.splitlines()
    current_row = 1
    table_number = 1
    sheet_count = 1
    table_count = 0
    heading_count = 0
    paragraph_count = 0
    first_sheet_named = False
    index = 0

    while index < len(lines):
        raw_line = lines[index]
        line = raw_line.strip()

        if not line:
            index += 1
            continue

        sheet_match = SHEET_HEADING_PATTERN.match(line)
        if sheet_match:
            sheet_name = sheet_match.group(1).strip() or f"Sheet{sheet_count}"
            if not first_sheet_named and current_row == 1 and worksheet.title == "Sheet1":
                worksheet.title = sheet_name[:31]
            else:
                worksheet = workbook.create_sheet(title=sheet_name[:31])
                current_row = 1
                sheet_count += 1
            first_sheet_named = True
            index += 1
            continue

        if line.startswith("|"):
            table_lines = []
            while index < len(lines) and lines[index].strip().startswith("|"):
                table_lines.append(lines[index].strip())
                index += 1
            parsed_rows = [split_table_row(item) for item in table_lines]
            if len(parsed_rows) >= 2 and is_separator_row(parsed_rows[1]):
                parsed_rows = [parsed_rows[0], *parsed_rows[2:]]
            if len(parsed_rows) >= 2:
                current_row, _ = write_table(worksheet, current_row, parsed_rows, table_number)
                table_number += 1
                table_count += 1
            continue

        if line.startswith("#"):
            level = len(line) - len(line.lstrip("#"))
            text = line.lstrip("#").strip()
            current_row = write_heading(worksheet, current_row, text, level)
            heading_count += 1
            index += 1
            continue

        current_row = write_paragraph(worksheet, current_row, raw_line.strip())
        paragraph_count += 1
        index += 1

    for sheet in workbook.worksheets:
        auto_fit_columns(sheet)

    stats = {
        "sheet_count": len(workbook.worksheets),
        "table_count": table_count,
        "heading_count": heading_count,
        "paragraph_count": paragraph_count,
    }
    return workbook, stats


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Convert Markdown headings and tables into an XLSX workbook.")
    parser.add_argument("input", nargs="?", help="Optional Markdown file path. If omitted, reads stdin unless --markdown is used.")
    parser.add_argument("--markdown", help="Inline Markdown content.")
    parser.add_argument("--output", required=True, help="Output XLSX path.")
    parser.add_argument("--title", help="Workbook title metadata.")
    parser.add_argument("--author", help="Workbook author metadata.")
    parser.add_argument("--subject", help="Workbook subject metadata.")
    parser.add_argument("--summary", action="store_true", help="Print a JSON summary after writing the workbook.")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    markdown = read_markdown(args.input, args.markdown)
    workbook, stats = build_workbook(markdown, title=args.title, author=args.author, subject=args.subject)
    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)

    if args.summary:
        print(
            json.dumps(
                {
                    "path": str(output_path.resolve()),
                    "stats": stats,
                },
                ensure_ascii=False,
                indent=2,
            )
        )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
